#!/usr/bin/env python3
import argparse
import datetime as dt
import json
import re
from pathlib import Path

import openpyxl
from docx import Document


def parse_sheet_date(sheet_name: str) -> dt.date | None:
    match = re.search(r"(\d{1,2})-(\d{1,2})(?:-(\d{2}))?", sheet_name)
    if not match:
        return None
    month = int(match.group(1))
    day = int(match.group(2))
    year = 2000 + int(match.group(3) or "26")
    try:
        return dt.date(year, month, day)
    except ValueError:
        return None


def iso_week(date_value: dt.date) -> str:
    iso = date_value.isocalendar()
    return f"{iso.year}-W{iso.week:02d}"


def to_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "")
        if cleaned == "":
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def parse_workbook(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    loads = []
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("Sheet"):
            continue
        day = parse_sheet_date(sheet_name)
        if not day:
            continue
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            drop_lot = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
            ref = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            line_haul = to_number(row[22] if len(row) > 22 else None)
            loaded_miles = to_number(row[23] if len(row) > 23 else None)
            pu_deadhead = to_number(row[24] if len(row) > 24 else None) or 0.0
            del_deadhead = to_number(row[25] if len(row) > 25 else None) or 0.0
            if not drop_lot or not ref or not line_haul or not loaded_miles:
                continue
            if line_haul <= 0 or loaded_miles <= 0:
                continue
            loads.append(
                {
                    "sheetName": sheet_name,
                    "sheetRow": row_idx,
                    "pickupDate": day.isoformat(),
                    "weekIso": iso_week(day),
                    "dropLot": drop_lot,
                    "ref": ref,
                    "statusText": str(row[2]).strip() if len(row) > 2 and row[2] is not None else "",
                    "issueText": str(row[3]).strip() if len(row) > 3 and row[3] is not None else "",
                    "routeId": str(row[5]).strip() if len(row) > 5 and row[5] is not None else "",
                    "loadNumber": str(row[6]).strip() if len(row) > 6 and row[6] is not None else "",
                    "pickupNumber": str(row[7]).strip() if len(row) > 7 and row[7] is not None else "",
                    "brokerName": str(row[8]).strip() if len(row) > 8 and row[8] is not None else "",
                    "driverAssignment": str(row[11]).strip() if len(row) > 11 and row[11] is not None else "",
                    "commodity": str(row[13]).strip() if len(row) > 13 and row[13] is not None else "",
                    "equipmentNeeds": str(row[14]).strip() if len(row) > 14 and row[14] is not None else "",
                    "shipperName": str(row[15]).strip() if len(row) > 15 and row[15] is not None else "",
                    "pickupCityState": str(row[16]).strip() if len(row) > 16 and row[16] is not None else "",
                    "pickupWindow": str(row[17]).strip() if len(row) > 17 and row[17] is not None else "",
                    "receiverName": str(row[18]).strip() if len(row) > 18 and row[18] is not None else "",
                    "deliveryCityState": str(row[19]).strip() if len(row) > 19 and row[19] is not None else "",
                    "deliveryWindow": str(row[20]).strip() if len(row) > 20 and row[20] is not None else "",
                    "lineHaulRate": round(line_haul, 4),
                    "loadedMiles": round(loaded_miles, 4),
                    "puDeadheadMiles": round(pu_deadhead, 4),
                    "delDeadheadMiles": round(del_deadhead, 4),
                }
            )
    return loads


def parse_doc(path: Path):
    doc = Document(path)
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    target_by_code = {}
    empty_pct_by_week = {}

    for text in paragraphs:
        if "Target:" in text and "$" in text:
            amount_match = re.search(r"Target:\s*\$([0-9,]+)", text)
            code_match = re.search(r"\(([A-Z0-9]{3,6})\)", text)
            if amount_match and code_match:
                target_by_code[code_match.group(1)] = amount_match.group(1).replace(",", "")

    for idx, text in enumerate(paragraphs):
        range_match = re.search(r"(\d{1,2})/(\d{1,2})/(\d{2})\s*-\s*(\d{1,2})/(\d{1,2})/(\d{2})", text)
        if not range_match:
            continue
        start = dt.date(2000 + int(range_match.group(3)), int(range_match.group(1)), int(range_match.group(2)))
        week = iso_week(start)
        next_line = paragraphs[idx + 1] if idx + 1 < len(paragraphs) else ""
        pct_match = re.search(r"empty mile pct\.\s*:\s*%?([0-9]+(?:\.[0-9]+)?)", next_line, re.IGNORECASE)
        if pct_match:
            empty_pct_by_week[week] = float(pct_match.group(1))

    return {
        "targetByCode": target_by_code,
        "emptyMilePctByWeek": empty_pct_by_week,
        "paragraphs": paragraphs,
    }


def main():
    parser = argparse.ArgumentParser(description="Parse blueprint spreadsheet and KPI doc into JSON payload.")
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--docx", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    loads = parse_workbook(Path(args.xlsx))
    doc_payload = parse_doc(Path(args.docx))
    weeks = sorted({row["weekIso"] for row in loads}.union(doc_payload["emptyMilePctByWeek"].keys()))

    output = {
        "loads": loads,
        "targetByCode": doc_payload["targetByCode"],
        "emptyMilePctByWeek": doc_payload["emptyMilePctByWeek"],
        "weeks": weeks,
    }
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(output, indent=2), encoding="utf-8")
    print(f"Parsed {len(loads)} load rows across {len(weeks)} week(s).")
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
